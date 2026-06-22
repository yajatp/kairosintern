"""Export helpers for Kairos lead data."""

from __future__ import annotations

import io

import pandas as pd


# Pain score fill colours (openpyxl ARGB, no leading #)
_SCORE_FILLS: list[tuple[float, float, str]] = [
    (6.0,  9.0,  "FFFCA5A5"),  # red    ≥ 6
    (4.0,  6.0,  "FFFDBA74"),  # orange 4–5
    (2.0,  4.0,  "FFFDE68A"),  # yellow 2–3
    (0.0,  2.0,  "FFBBF7D0"),  # green  0–1
]


def _score_fill(score: float):
    """Return an openpyxl PatternFill for the given pain score."""
    from openpyxl.styles import PatternFill

    for lo, hi, argb in _SCORE_FILLS:
        if score >= lo:
            return PatternFill(fill_type="solid", fgColor=argb)
    return PatternFill(fill_type="solid", fgColor="FFBBF7D0")


def df_to_xlsx_bytes(df: pd.DataFrame, sheet_name: str = "Leads") -> bytes:
    """Convert a DataFrame to an XLSX byte string with formatting.

    Features:
    - Frozen top row
    - Bold headers
    - Auto-width columns (capped at 60)
    - Pain Score column colour-coded by value
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    headers = list(df.columns)
    ws.append(headers)

    # Bold headers + freeze top row
    for cell in ws[1]:
        cell.font = Font(bold=True)
    ws.freeze_panes = "A2"

    # Find Pain Score column index (1-based)
    pain_col_idx: int | None = None
    if "Pain Score" in headers:
        pain_col_idx = headers.index("Pain Score") + 1  # 1-based

    # Write data rows
    for _, row in df.iterrows():
        ws.append([row[col] for col in headers])

    # Colour Pain Score cells
    if pain_col_idx is not None:
        col_letter = get_column_letter(pain_col_idx)
        for row_idx in range(2, ws.max_row + 1):
            cell = ws[f"{col_letter}{row_idx}"]
            try:
                score = float(cell.value or 0)
            except (TypeError, ValueError):
                score = 0.0
            cell.fill = _score_fill(score)

    # Auto-width (estimate from max string length in each column)
    for col_idx, col_cells in enumerate(ws.iter_cols(min_row=1), start=1):
        max_len = 0
        for cell in col_cells:
            try:
                cell_len = len(str(cell.value)) if cell.value is not None else 0
            except Exception:
                cell_len = 0
            if cell_len > max_len:
                max_len = cell_len
        adjusted = min(max_len + 2, 60)
        ws.column_dimensions[get_column_letter(col_idx)].width = adjusted

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
